import openpyxl

# file -> workbook -> sheets -> rows -> columns

file = "D:\\Selenium\\Excel files\\demo_sheet.xlsx"  # file path
workbook = openpyxl.load_workbook(file)  # load the workbook
sheet = workbook["Sheet1"]

rows = sheet.max_row  # number of rows present in sheet
cols = sheet.max_column  # number of columns present in sheet

print("Number of rows:", rows)
print("Number or columns", cols)

# reading all the data from rows and columns
for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(r, c).value, end="  --->  ")
    print()
