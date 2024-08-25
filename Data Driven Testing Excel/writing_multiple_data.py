import openpyxl

file = "D:\\Selenium\\Excel files\\empty_sheet1.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

sheet.cell(1, 1).value = "ID"
sheet.cell(1, 2).value = "NAME"
sheet.cell(1, 3).value = "EMAIL"

sheet.cell(2, 1).value = "x001"
sheet.cell(2, 2).value = "Nilanchala Samal"
sheet.cell(2, 3).value = "nilanchal960@gmail.com"

sheet.cell(3, 1).value = "x002"
sheet.cell(3, 2).value = "Biswajit Bhanja"
sheet.cell(3, 3).value = "biswajit1999@gmail.com"

sheet.cell(4, 1).value = "x003"
sheet.cell(4, 2).value = "Surya Prakash Ojha"
sheet.cell(4, 3).value = "suryaojha@gmail.com"

workbook.save(file)
